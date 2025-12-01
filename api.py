#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
1688 ä»¥å›¾æœè´§ API æœåŠ¡

æ”¯æŒ:
- å•å¼ å›¾ç‰‡æœç´¢
- æ‰¹é‡å›¾ç‰‡æœç´¢ï¼ˆæœ€å¤š 3000 å¼ ï¼‰
- å¤„ç†å®Œæˆåå‘é€ Excel åˆ°é‚®ç®±
"""

import os
import uuid
import asyncio
import tempfile
from typing import List, Optional, Dict
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.header import Header
from email import encoders
from io import BytesIO

from fastapi import FastAPI, UploadFile, File, HTTPException, Query, BackgroundTasks
from pydantic import BaseModel
import requests
import aiosmtplib
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from lib.ali1688 import ali1688
from lib.ali1688.search import fetch_product_links_async, get_search_url
from lib.proxy import get_new_proxy, ProxyInfo
from config.email_config import SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASS, SENDER_NAME


app = FastAPI(
    title="1688 ä»¥å›¾æœè´§ API",
    description="ä¸Šä¼ å›¾ç‰‡ï¼Œè·å– 1688 ç›¸ä¼¼äº§å“é“¾æ¥ã€‚æ”¯æŒæ‰¹é‡å¤„ç† + é‚®ä»¶é€šçŸ¥ã€‚",
    version="2.0.0"
)


# ============== ä»»åŠ¡å­˜å‚¨ ==============
tasks_store: Dict[str, dict] = {}


# ============== æ•°æ®æ¨¡å‹ ==============

class Product(BaseModel):
    title: str
    url: str
    offer_id: str


class SearchResponse(BaseModel):
    success: bool
    image_id: Optional[str] = None
    search_url: Optional[str] = None
    products: List[Product] = []
    error: Optional[str] = None


class EmailBatchRequest(BaseModel):
    image_urls: List[str]
    email: str
    limit: int = 5


# ============== æ ¸å¿ƒå‡½æ•° ==============

def upload_image_to_1688(image_path: str) -> dict:
    """ä¸Šä¼ å›¾ç‰‡åˆ° 1688"""
    upload = ali1688.Ali1688Upload()
    res = upload.upload(filename=image_path)
    return res.json()


async def search_products(
    image_path: str, 
    limit: int = 5,
    proxy_info: Optional[ProxyInfo] = None,  # å¯æŒ‡å®šä»£ç†
    max_retries: int = 3  # æœ€å¤§é‡è¯•æ¬¡æ•°
) -> SearchResponse:
    """æœç´¢äº§å“ï¼ˆå¸¦é‡è¯•æœºåˆ¶ï¼‰"""
    import asyncio
    
    last_error = None
    
    for attempt in range(max_retries):
        try:
            # ä¸Šä¼ å›¾ç‰‡åˆ° 1688
            data = upload_image_to_1688(image_path)
            
            # æ£€æŸ¥ä¸Šä¼ ç»“æœ
            ret_msg = data.get("ret", [""])[0] if data.get("ret") else ""
            if ret_msg != "SUCCESS::è°ƒç”¨æˆåŠŸ":
                last_error = "ä¸Šä¼ å¤±è´¥"
                if attempt < max_retries - 1:
                    wait_time = (attempt + 1) * 2
                    print(f"âš ï¸ ä¸Šä¼ å¤±è´¥ï¼Œ{wait_time}ç§’åé‡è¯•... ({attempt + 1}/{max_retries})")
                    await asyncio.sleep(wait_time)
                    continue
                return SearchResponse(success=False, error=last_error)
            
            image_id = data.get("data", {}).get("imageId", "")
            if not image_id:
                return SearchResponse(success=False, error="æœªè·å–åˆ° imageId")
            
            search_url = get_search_url(image_id)
            products_data = await fetch_product_links_async(
                image_id, 
                limit=limit, 
                headless=True,
                proxy_info=proxy_info  # ä¼ é€’æŒ‡å®šçš„ä»£ç†
            )
            products = [Product(**p) for p in products_data]
            
            return SearchResponse(
                success=True,
                image_id=image_id,
                search_url=search_url,
                products=products
            )
            
        except (TimeoutError, ConnectionError) as e:
            last_error = str(e)
            if attempt < max_retries - 1:
                wait_time = (attempt + 1) * 2
                print(f"âš ï¸ è¿æ¥è¶…æ—¶: {e}ï¼Œ{wait_time}ç§’åé‡è¯•... ({attempt + 1}/{max_retries})")
                await asyncio.sleep(wait_time)
                continue
            return SearchResponse(success=False, error=f"è¿æ¥è¶…æ—¶: {last_error}")
            
        except Exception as e:
            error_msg = str(e)
            # æ£€æŸ¥æ˜¯å¦æ˜¯å¯é‡è¯•çš„é”™è¯¯
            if any(keyword in error_msg for keyword in ["timeout", "timed out", "Connection aborted", "Connection reset"]):
                last_error = error_msg
                if attempt < max_retries - 1:
                    wait_time = (attempt + 1) * 2
                    print(f"âš ï¸ {error_msg[:50]}...ï¼Œ{wait_time}ç§’åé‡è¯•... ({attempt + 1}/{max_retries})")
                    await asyncio.sleep(wait_time)
                    continue
            return SearchResponse(success=False, error=error_msg)
    
    return SearchResponse(success=False, error=f"é‡è¯• {max_retries} æ¬¡åä»å¤±è´¥: {last_error}")


# ============== Excel ç”Ÿæˆ ==============

def create_excel(results: List[dict]) -> BytesIO:
    """
    ç”Ÿæˆ Excel æ–‡ä»¶
    è¡¨å¤´ï¼šåºå·ã€åŸå›¾URLã€çŠ¶æ€ã€äº§å“é“¾æ¥
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "æœç´¢ç»“æœ"
    
    # è¡¨å¤´
    headers = ["åºå·", "åŸå›¾URL", "çŠ¶æ€", "äº§å“é“¾æ¥"]
    ws.append(headers)
    
    # è®¾ç½®è¡¨å¤´æ ·å¼
    for col in range(1, 5):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # æ•°æ®è¡Œ
    for idx, result in enumerate(results, 1):
        row_num = idx + 1
        
        # åºå·
        ws.cell(row=row_num, column=1, value=idx)
        
        # åŸå›¾ URL
        ws.cell(row=row_num, column=2, value=result.get("image_url", ""))
        
        # çŠ¶æ€
        if result.get("success"):
            ws.cell(row=row_num, column=3, value="æˆåŠŸ")
            
            # äº§å“é“¾æ¥ï¼ˆ5 ä¸ªé“¾æ¥ç”¨é€—å·åˆ†éš”ï¼‰
            products = result.get("products", [])
            links = [p.get("url", "") for p in products[:5]]
            ws.cell(row=row_num, column=4, value=",".join(links))
        else:
            ws.cell(row=row_num, column=3, value="å¤±è´¥: " + result.get("error", "æœªçŸ¥é”™è¯¯"))
            ws.cell(row=row_num, column=4, value="")
    
    # è°ƒæ•´åˆ—å®½
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 150
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ============== é‚®ä»¶å‘é€ ==============

async def send_email_with_excel(to_email: str, excel_buffer: BytesIO, task_info: dict):
    """å‘é€å¸¦ Excel é™„ä»¶çš„é‚®ä»¶"""
    msg = MIMEMultipart()
    # RFC2047 ç¼–ç ä¸­æ–‡å‘ä»¶äººåç§°
    from email.utils import formataddr
    msg["From"] = formataddr((str(Header(SENDER_NAME, "utf-8")), SMTP_USER))
    msg["To"] = to_email
    msg["Subject"] = Header(f"1688 ä»¥å›¾æœè´§ç»“æœ - {task_info['total']} å¼ å›¾ç‰‡", "utf-8")
    
    # é‚®ä»¶æ­£æ–‡
    body = f"""æ‚¨å¥½ï¼

æ‚¨æäº¤çš„ 1688 ä»¥å›¾æœè´§ä»»åŠ¡å·²å®Œæˆã€‚

ğŸ“‹ ä»»åŠ¡ä¿¡æ¯ï¼š
â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”
â€¢ ä»»åŠ¡ ID: {task_info['task_id']}
â€¢ å›¾ç‰‡æ•°é‡: {task_info['total']}
â€¢ æˆåŠŸæ•°é‡: {task_info['success_count']}
â€¢ å¤±è´¥æ•°é‡: {task_info['fail_count']}
â€¢ å¤„ç†æ—¶é—´: {task_info['duration']} ç§’
â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”

è¯·æŸ¥çœ‹é™„ä»¶ä¸­çš„ Excel æ–‡ä»¶è·å–è¯¦ç»†ç»“æœã€‚

Excel è¡¨æ ¼è¯´æ˜ï¼š
â€¢ åºå·ï¼šå›¾ç‰‡åºå·
â€¢ åŸå›¾URLï¼šæ‚¨æäº¤çš„å›¾ç‰‡åœ°å€
â€¢ çŠ¶æ€ï¼šæˆåŠŸ/å¤±è´¥
â€¢ äº§å“é“¾æ¥ï¼š5 ä¸ªç›¸ä¼¼äº§å“çš„é“¾æ¥ï¼ˆé€—å·åˆ†éš”ï¼‰

--
1688 ä»¥å›¾æœè´§ API
"""
    msg.attach(MIMEText(body, "plain", "utf-8"))
    
    # Excel é™„ä»¶
    attachment = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    attachment.set_payload(excel_buffer.getvalue())
    encoders.encode_base64(attachment)
    filename = f"1688_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    attachment.add_header("Content-Disposition", "attachment", filename=filename)
    msg.attach(attachment)
    
    # å‘é€é‚®ä»¶ - QQ é‚®ç®± 465 ç«¯å£ä½¿ç”¨ SSL
    import ssl
    context = ssl.create_default_context()
    
    async with aiosmtplib.SMTP(
        hostname=SMTP_HOST, 
        port=SMTP_PORT, 
        use_tls=True,
        tls_context=context
    ) as smtp:
        await smtp.login(SMTP_USER, SMTP_PASS)
        await smtp.send_message(msg)


# ============== åå°ä»»åŠ¡å¤„ç† ==============

async def process_email_batch_task(task_id: str, image_urls: List[str], email: str, limit: int):
    """åå°å¤„ç†æ‰¹é‡ä»»åŠ¡å¹¶å‘é€é‚®ä»¶"""
    task = tasks_store[task_id]
    task["status"] = "processing"
    start_time = datetime.now()
    
    results = []
    results_lock = asyncio.Lock()
    
    # ğŸš€ å¹¶è¡Œå·¥ä½œé˜Ÿåˆ—ï¼š3 ä¸ª worker å¹¶è¡Œï¼Œæ¯ä¸ª worker å†…éƒ¨ä¸²è¡Œï¼ˆè·å–ä»£ç†â†’ä½¿ç”¨â†’è·å–â†’ä½¿ç”¨ï¼‰
    WORKER_COUNT = 3
    queue = asyncio.Queue()
    
    # å¡«å……ä»»åŠ¡é˜Ÿåˆ—
    for i, url in enumerate(image_urls):
        await queue.put((i, url))
    
    async def worker(worker_id: int):
        """å·¥ä½œåç¨‹ï¼šä»é˜Ÿåˆ—å–ä»»åŠ¡ï¼Œè·å–ä»£ç†ï¼Œæ‰§è¡Œæœç´¢"""
        while True:
            try:
                index, url = queue.get_nowait()
            except asyncio.QueueEmpty:
                break
            
            result = {"image_url": url, "index": index}
            try:
                # ä¸‹è½½å›¾ç‰‡
                response = requests.get(url, timeout=30)
                response.raise_for_status()
                
                with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp:
                    tmp.write(response.content)
                    tmp_path = tmp.name
                
                try:
                    # ğŸ”„ è·å–ä»£ç†å¹¶ç«‹å³ä½¿ç”¨ï¼ˆworker å†…ä¸²è¡Œï¼Œä¸ä¼šç©ºç­‰ï¼‰
                    proxy = get_new_proxy()
                    print(f"[Worker-{worker_id}] ğŸŒ ä½¿ç”¨ä»£ç†: {proxy.server if proxy else 'None'}")
                    search_result = await search_products(tmp_path, limit=limit, proxy_info=proxy)
                finally:
                    if os.path.exists(tmp_path):
                        os.unlink(tmp_path)
                
                result.update({
                    "success": search_result.success,
                    "products": [p.dict() for p in search_result.products] if search_result.products else [],
                    "error": search_result.error
                })
            except Exception as e:
                result.update({"success": False, "error": str(e), "products": []})
            
            async with results_lock:
                results.append(result)
                task["completed"] += 1
                print(f"[{task_id[:8]}] è¿›åº¦: {task['completed']}/{task['total']}")
    
    print(f"[{task_id[:8]}] å¼€å§‹å¤„ç† {len(image_urls)} å¼ å›¾ç‰‡ï¼ˆ{WORKER_COUNT} ä¸ª worker å¹¶è¡Œï¼‰")
    
    # å¯åŠ¨ worker å¹¶ç­‰å¾…å®Œæˆ
    workers = [worker(i) for i in range(WORKER_COUNT)]
    await asyncio.gather(*workers)
    
    # æŒ‰åŸå§‹é¡ºåºæ’åº
    results.sort(key=lambda x: x.get("index", 0))
    
    # ç»Ÿè®¡
    success_count = sum(1 for r in results if r.get("success"))
    fail_count = len(results) - success_count
    duration = (datetime.now() - start_time).total_seconds()
    
    task_info = {
        "task_id": task_id,
        "total": len(results),
        "success_count": success_count,
        "fail_count": fail_count,
        "duration": round(duration, 1)
    }
    
    # ç”Ÿæˆ Excel
    print(f"[{task_id[:8]}] ç”Ÿæˆ Excel...")
    excel_buffer = create_excel(results)
    
    # å‘é€é‚®ä»¶
    try:
        print(f"[{task_id[:8]}] å‘é€é‚®ä»¶åˆ° {email}...")
        await send_email_with_excel(email, excel_buffer, task_info)
        task["status"] = "completed"
        task["message"] = f"ç»“æœå·²å‘é€åˆ° {email}"
        print(f"[{task_id[:8]}] âœ… å®Œæˆï¼é‚®ä»¶å·²å‘é€")
    except Exception as e:
        task["status"] = "email_failed"
        task["message"] = f"å¤„ç†å®Œæˆä½†é‚®ä»¶å‘é€å¤±è´¥: {e}"
        print(f"[{task_id[:8]}] âŒ é‚®ä»¶å‘é€å¤±è´¥: {e}")
    
    task["completed"] = len(results)
    task["success_count"] = success_count
    task["fail_count"] = fail_count
    task["duration"] = round(duration, 1)


# ============== API ç«¯ç‚¹ ==============

@app.get("/")
async def root():
    """API æ ¹è·¯å¾„"""
    return {
        "name": "1688 ä»¥å›¾æœè´§ API",
        "version": "2.0.0",
        "endpoints": {
            "POST /search/upload": "å•å¼ å›¾ç‰‡æœç´¢",
            "GET /search/url": "é€šè¿‡ URL æœç´¢",
            "POST /batch/email": "æ‰¹é‡æœç´¢ + é‚®ä»¶é€šçŸ¥ï¼ˆæœ€å¤š 3000 å¼ ï¼‰",
            "GET /batch/status/{task_id}": "æŸ¥è¯¢ä»»åŠ¡çŠ¶æ€"
        }
    }


@app.post("/search/upload", response_model=SearchResponse)
async def search_by_upload(
    file: UploadFile = File(..., description="è¦æœç´¢çš„å›¾ç‰‡æ–‡ä»¶"),
    limit: int = Query(5, ge=1, le=20, description="è¿”å›äº§å“æ•°é‡")
):
    """å•å¼ å›¾ç‰‡æœç´¢"""
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="è¯·ä¸Šä¼ å›¾ç‰‡æ–‡ä»¶")
    
    suffix = os.path.splitext(file.filename or "image.jpg")[1] or ".jpg"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name
    
    try:
        result = await search_products(tmp_path, limit=limit)
        return result
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


@app.get("/search/url", response_model=SearchResponse)
async def search_by_url(
    image_url: str = Query(..., description="å›¾ç‰‡ URL"),
    limit: int = Query(5, ge=1, le=20, description="è¿”å›äº§å“æ•°é‡")
):
    """é€šè¿‡ URL æœç´¢"""
    try:
        response = requests.get(image_url, timeout=30)
        response.raise_for_status()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"ä¸‹è½½å›¾ç‰‡å¤±è´¥: {e}")
    
    with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp:
        tmp.write(response.content)
        tmp_path = tmp.name
    
    try:
        result = await search_products(tmp_path, limit=limit)
        return result
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


@app.post("/batch/email")
async def batch_search_with_email(
    request: EmailBatchRequest,
    background_tasks: BackgroundTasks
):
    """
    æ‰¹é‡æœç´¢ + é‚®ä»¶é€šçŸ¥
    
    - **image_urls**: å›¾ç‰‡ URL åˆ—è¡¨ï¼ˆæœ€å¤š 3000 ä¸ªï¼‰
    - **email**: ç»“æœå‘é€åˆ°çš„é‚®ç®±åœ°å€
    - **limit**: æ¯å¼ å›¾ç‰‡è¿”å›çš„äº§å“æ•°é‡ï¼ˆé»˜è®¤ 5ï¼‰
    
    ğŸ’¡ æäº¤åç«‹å³è¿”å›ï¼Œå¤„ç†å®Œæˆåå‘é€ Excel åˆ°é‚®ç®±
    
    Excel æ ¼å¼ï¼šåºå·ã€åŸå›¾URLã€çŠ¶æ€ã€äº§å“é“¾æ¥ï¼ˆ5ä¸ªé“¾æ¥ç”¨é€—å·åˆ†éš”ï¼‰
    """
    if len(request.image_urls) > 3000:
        raise HTTPException(status_code=400, detail="æœ€å¤šæ”¯æŒ 3000 ä¸ª URL")
    
    if len(request.image_urls) == 0:
        raise HTTPException(status_code=400, detail="è¯·æä¾›è‡³å°‘ 1 ä¸ª URL")
    
    if not request.email or "@" not in request.email:
        raise HTTPException(status_code=400, detail="è¯·æä¾›æœ‰æ•ˆçš„é‚®ç®±åœ°å€")
    
    # åˆ›å»ºä»»åŠ¡
    task_id = str(uuid.uuid4())
    tasks_store[task_id] = {
        "task_id": task_id,
        "status": "pending",
        "total": len(request.image_urls),
        "completed": 0,
        "success_count": 0,
        "fail_count": 0,
        "email": request.email,
        "created_at": datetime.now().isoformat(),
        "message": ""
    }
    
    # åå°å¤„ç†
    background_tasks.add_task(
        process_email_batch_task, 
        task_id, 
        request.image_urls, 
        request.email, 
        request.limit
    )
    
    # é¢„ä¼°æ—¶é—´ï¼ˆæ¯å¼ çº¦ 15-30 ç§’ï¼Œ3 å¹¶å‘ï¼‰
    estimated_minutes = (len(request.image_urls) * 20) // 60 // 3
    
    return {
        "task_id": task_id,
        "status": "pending",
        "total": len(request.image_urls),
        "email": request.email,
        "message": f"ä»»åŠ¡å·²æäº¤ï¼å¤„ç†å®Œæˆåç»“æœå°†å‘é€åˆ° {request.email}",
        "estimated_time": f"çº¦ {estimated_minutes} - {estimated_minutes * 2} åˆ†é’Ÿ"
    }


@app.get("/batch/status/{task_id}")
async def get_task_status(task_id: str):
    """æŸ¥è¯¢ä»»åŠ¡çŠ¶æ€"""
    if task_id not in tasks_store:
        raise HTTPException(status_code=404, detail="ä»»åŠ¡ä¸å­˜åœ¨")
    
    task = tasks_store[task_id]
    return {
        "task_id": task["task_id"],
        "status": task["status"],
        "total": task["total"],
        "completed": task["completed"],
        "success_count": task.get("success_count", 0),
        "fail_count": task.get("fail_count", 0),
        "progress": f"{task['completed']}/{task['total']}",
        "email": task["email"],
        "message": task.get("message", ""),
        "duration": task.get("duration"),
        "created_at": task["created_at"]
    }


@app.get("/health")
async def health_check():
    """å¥åº·æ£€æŸ¥"""
    return {
        "status": "healthy",
        "active_tasks": len([t for t in tasks_store.values() if t["status"] == "processing"])
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
